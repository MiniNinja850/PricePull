#make a script to go get the price, time stamp, name of items. 
#use seleium, but don't use the auto chromedriver installer. 
#Due by next week.  
# or set variables
#print date item name and price
# https://www.lowes.com/pd/ANZZI-Bank-Series-65-in-Acrylic-Freestanding-Bathtub-with-Deck-Mounted-Faucet-in-White/5001949753
# https://www.homedepot.com/p/ANZZI-Bank-65-in-Acrylic-Flatbottom-Non-Whirlpool-Bathtub-with-Deck-Mounted-Faucet-in-White-FT-FR112473CH/313605873
# https://www.lowes.com/pd/AKDY-63-in-Brush-Stainless-Steel-2-Spray-Shower-Panel-System-Valve-Included/1000790644
# https://www.homedepot.com/p/AKDY-63-in-2-Massage-Spray-Shower-Panel-System-in-Stainless-Steel-with-Rainfall-Shower-Head-and-Shower-Wand-SP0107/304238187
import time
import os
import sys
import re
from xml.etree.ElementTree import XMLParser
import openpyxl
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService  
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

#Calls webdriver...
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

#Sets frame for window to open. 
chrome_options = Options()
chrome_options.add_argument("--window-size=1920,1080")
chrome_options.add_argument("--start-maximized")

#Website to pull info from: 
web_url1 = 'https://www.lowes.com/pd/ANZZI-Bank-Series-65-in-Acrylic-Freestanding-Bathtub-with-Deck-Mounted-Faucet-in-White/5001949753'
web_url2 ='https://www.homedepot.com/p/ANZZI-Bank-65-in-Acrylic-Flatbottom-Non-Whirlpool-Bathtub-with-Deck-Mounted-Faucet-in-White-FT-FR112473CH/313605873'
web_url3 = 'https://www.lowes.com/pd/AKDY-63-in-Brush-Stainless-Steel-2-Spray-Shower-Panel-System-Valve-Included/1000790644'
web_url4 = 'https://www.homedepot.com/p/AKDY-63-in-2-Massage-Spray-Shower-Panel-System-in-Stainless-Steel-with-Rainfall-Shower-Head-and-Shower-Wand-SP0107/304238187'

#XPATHS: # Xpath should look like: //span[@class='regular-price text-orange']/span[@class='price'] 
price1XP='//span[@class="item-price-dollar"]'
title1XP='//h1[@class="styles__H1-sc-11vpuyu-0 kQJGef typography variant--h1 align--left product-desc"]'
#'//h1[@class="styles__H1-sc-11vpuyu-0 kQJGef typography variant--h1 align--left product-desc"]'


price2XP='//div[@class="price-format__large price-format__main-price"]//span[2]'
# '//div[@class="price-format__large price-format__main-price"]'//span[2]
title2XP='//h1[@class="product-details__title"]'

price3XP='//span[@class="item-price-dollar"]'
title3XP='//h1[@class="styles__H1-sc-11vpuyu-0 kQJGef typography variant--h1 align--left product-desc"]'
# '//h1[@class="styles__H1-sc-11vpuyu-0 kQJGef typography variant--h1 align--left product-desc"]' color="text_primary">AKDY<!-- --> &nbsp;<!-- -->Brush Stainless Steel 2-Spray Shower Panel System</h1>

price4XP='//div[@class="price-format__large price-format__main-price"]//span[2]'
title4XP='//h1[@class="product-details__title"]'

# web driver goes to page
driver.get(web_url1)
driver.implicitly_wait(10)
time.sleep(5)
price1XP = driver.find_element(By.XPATH, price1XP).text

time.sleep(5)
title1XP = driver.find_element(By.XPATH, title1XP).text
time.sleep(2)

driver.get(web_url2)
time.sleep(3)
title2XP = driver.find_element(By.XPATH, title2XP).text
time.sleep(3)
price2XP = driver.find_element(By.XPATH, price2XP).text
time.sleep(2)

driver.get(web_url3)
time.sleep(3)
title3XP = driver.find_element(By.XPATH, title3XP).text
time.sleep(3)
price3XP = driver.find_element(By.XPATH, price3XP).text
time.sleep(2)

driver.get(web_url4)
time.sleep(3)
title4XP = driver.find_element(By.XPATH, title4XP).text
time.sleep(3)
price4XP = driver.find_element(By.XPATH, price4XP).text
time.sleep(2)

#End results: 
print(title1XP + price1XP)
print(title2XP + price2XP)
print(title3XP + price3XP)
print(title4XP + price4XP)

#workbook where the data will pull into: 
outputWB = openpyxl.Workbook("Data_Pull.xlsx") 
sheet2write = outputWB.active
sheet2write.cell(row=1, column=1).value = "Item1"
sheet2write.cell(row=1, column=2).value = "Price1"
sheet2write.cell(row=1, column=3).value = "Item2"
sheet2write.cell(row=1, column=4).value = "Price2"
sheet2write.cell(row=1, column=5).value = "Item3"
sheet2write.cell(row=1, column=6).value = "Price3"
sheet2write.cell(row=1, column=7).value = "Item4"
sheet2write.cell(row=1, column=8).value = "Price4"
outputWB.save("Data_Pull.xlsx") 

# Script is telling you that its complete: 
print('Scraper Process Completed')
timestamp=datetime.now()
tlog=timestamp.strftime("%m/%d/%Y %H:%M:%S")   
logp= open('Log.txt','a')
logp.write(tlog + ' - Scraper Process Completed' + ' \n \n')
logp.close()